from flask import Flask, render_template, request
import PyPDF2
from openpyxl import Workbook
from io import BytesIO
from flask import send_file
app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html')



@app.route('/process', methods=['POST'])
def process():
    uploaded_file = request.files['file']
    column_names = request.form['columns']

    # Save uploaded PDF to BytesIO
    pdf_data = BytesIO(uploaded_file.read())

    # Your PDF scraping code here...
    input_pdf = PyPDF2.PdfReader(pdf_data)

    main_list = column_names.split(',')

    wb = Workbook()
    ws = wb.active
    row_num = 1
    column_num = 1

    for i in range(len(main_list)):
        field = main_list[i]
        ws.cell(row=row_num, column=column_num, value=field)
        column_num += 1

    total_pages = len(input_pdf.pages)
    row_num = 2

    for i in range(total_pages):
        page = input_pdf.pages[i]
        page_content = page.extract_text()
        column_num = 1

        for i in range(len(main_list)):
            field = main_list[i]
            next_field = main_list[i + 1] if i + 1 < len(main_list) else None

            # Find position of fields from extracted text of PDF file
            field_pos = page_content.find(field)
            next_field_pos = page_content.find(next_field) if next_field else None

            # Find position of field values from extracted text of PDF file
            field_value_start_pos = field_pos + len(field)
            field_value_end_pos = next_field_pos if next_field_pos else None

            # Extract field values
            field_value = page_content[field_value_start_pos:field_value_end_pos]

            # Write field values into Excel
            ws.cell(row=row_num, column=column_num, value=field_value)
            column_num += 1

        row_num += 1

    # Save the Excel file to BytesIO
    excel_data = BytesIO()
    wb.save(excel_data)
    excel_data.seek(0)

    # Return the Excel file for download
    return send_file(excel_data, as_attachment=True, download_name='result.xlsx')

if __name__ == '__main__':
    app.run(debug=True)
